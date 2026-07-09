"""Draft-prep tool. Prepares cold outreach for a human to read and paste.

    python main.py --prep                # build drafts.md + queue.xlsx
    python main.py --prep --limit 10     # only draft the 10 best-ordered leads
    python main.py --mark-sent <ident>   # record what you actually sent
    python main.py --suppress <ident>    # permanent opt-out, all channels
    python main.py --stats               # what's been sent so far

THERE IS NO --send FLAG, AND THERE MUST NEVER BE ONE.

At ~25 usable leads, automation is the wrong build. Sending by hand is faster
than the automation that would replace it; Gmail cannot tell the difference
(deliverability is decided by domain reputation, authentication, bounce rate and
engagement, not by who pressed the button); and a human in the loop is the only
thing that stops `Hi {{FirstName}},` going out 25 times, which is unrecoverable.

Send ~10-15 emails a day from ONE mailbox. That IS the warmup — no paid warmup
tool is needed. Spreading 25 emails across 5 mailboxes gives all five of them no
reputation at all.
"""

import argparse
import sys

import config
import draft
import leads as leads_mod
import verify
from ledger import Ledger


def _fmt_lead(lead) -> str:
    tag = "role" if lead.channel == config.CHANNEL_EMAIL and lead.is_role_address else ""
    stale = f"{lead.stale_days}d stale" if lead.stale_days else ""
    bits = [b for b in (lead.category, stale, tag) if b]
    return f"{lead.business} ({', '.join(bits)})" if bits else lead.business


def cmd_prep(limit: int) -> int:
    all_leads, rejected = leads_mod.load_all()
    print(f"[leads] loaded {len(all_leads)} leads, "
          f"rejected {len(rejected)} before drafting")
    for identifier, reason in rejected:
        print(f"[purge] {identifier}: {reason}")

    ledger = Ledger()
    ledger.clear_drafts()

    ordered = verify.send_order(all_leads)
    drafted, held, skipped = [], [], []

    for i, lead in enumerate(ordered):
        if ledger.is_suppressed(lead.channel, lead.identifier):
            skipped.append((lead, "opted out"))
            continue
        if ledger.already_sent(lead.channel, lead.identifier):
            skipped.append((lead, "already sent"))
            continue
        hit = ledger.domain_contacted(lead.domain)
        if hit and hit != lead.identifier.lower():
            skipped.append((lead, f"already messaged {hit} at this company"))
            continue

        verdict, note = verify.check(lead)
        if verdict == verify.DROP:
            skipped.append((lead, note))
            continue
        if verdict == verify.HOLD:
            held.append((lead, note))
            continue

        if limit and len(drafted) >= limit:
            skipped.append((lead, f"beyond --limit {limit}"))
            continue

        try:
            subject, body = draft.build(lead, ledger, attempt_order=i)
        except draft.DraftError as exc:
            reason = str(exc)
            if lead.channel == config.CHANNEL_EMAIL and draft.is_opinion_only(lead):
                reason = "audit found only opinions, no checkable fact"
            skipped.append((lead, reason))
            continue

        ledger.hold_pending(lead.identifier, body)
        ledger.save_draft(lead.channel, lead.identifier, subject, body,
                          lead.business, lead.domain)
        drafted.append((lead, subject, body))

    _write_drafts_md(drafted, held, skipped)
    _write_queue_xlsx(drafted, held, skipped)

    print(f"\n[prep] {len(drafted)} drafted, {len(held)} held, {len(skipped)} skipped")
    print(f"[prep] drafts -> {config.DRAFTS_FILE}")
    print(f"[prep] queue  -> {config.QUEUE_XLSX}")
    print("\nRead every draft before sending. Send the freshest domains first,")
    print("~10/day, from one mailbox. Any bounce: stop and re-check the HOLD list.")
    ledger.close()
    return 0


def _write_drafts_md(drafted, held, skipped) -> None:
    lines = ["# Outreach drafts", "",
             "Read each one. Fix anything that reads wrong. Paste it yourself.",
             "Then run `python main.py --mark-sent <identifier>`.", ""]

    for channel, title in ((config.CHANNEL_EMAIL, "Emails"),
                           (config.CHANNEL_INSTAGRAM, "Instagram DMs")):
        rows = [d for d in drafted if d[0].channel == channel]
        if not rows:
            continue
        lines += [f"## {title} ({len(rows)})", ""]
        for lead, subject, body in rows:
            lines.append(f"### {lead.business}")
            lines.append(f"- **To:** `{lead.identifier}`")
            if subject:
                lines.append(f"- **Subject:** {subject}")
            if lead.stale_days:
                lines.append(f"- **Site last touched:** {lead.stale_days} days ago")
            if lead.channel == config.CHANNEL_EMAIL and lead.is_role_address:
                lines.append("- **Note:** role address — expect a low reply rate")
            lines += ["", "```text", body, "```", ""]

    if held:
        lines += ["## Held back", "",
                  "Not drafted on purpose. Send these only once the mailbox has a "
                  "reputation.", ""]
        lines += [f"- `{l.identifier}` — {note}" for l, note in held] + [""]
    if skipped:
        lines += ["## Skipped", ""]
        lines += [f"- `{l.identifier}` — {note}" for l, note in skipped] + [""]

    with open(config.DRAFTS_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def _write_queue_xlsx(drafted, held, skipped) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[prep] openpyxl not installed — skipping queue.xlsx")
        return

    columns = ["Status", "Channel", "Identifier", "Business", "Category",
               "Stale Days", "Subject", "Note"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Queue"
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for lead, subject, _ in drafted:
        ws.append(["DRAFTED", lead.channel, lead.identifier, lead.business,
                   lead.category, lead.stale_days or "", subject, ""])
    for lead, note in held:
        ws.append(["HELD", lead.channel, lead.identifier, lead.business,
                   lead.category, lead.stale_days or "", "", note])
    for lead, note in skipped:
        ws.append(["SKIPPED", lead.channel, lead.identifier, lead.business,
                   lead.category, lead.stale_days or "", "", note])

    for i, col in enumerate(columns, 1):
        width = max([len(col)] + [len(str(c.value or ""))
                                  for c in ws[get_column_letter(i)]])
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 60)
    ws.freeze_panes = "A2"
    wb.save(config.QUEUE_XLSX)


def cmd_mark_sent(identifier: str) -> int:
    ledger = Ledger()
    row = ledger.get_draft(identifier)
    if row is None:
        print(f"[ledger] no pending draft for '{identifier}'. Run --prep first, "
              "or check the spelling.")
        ledger.close()
        return 1
    ok = ledger.mark_sent(row["channel"], row["identifier"], row["body"],
                          row["business"], row["domain"], row["subject"])
    print(f"[ledger] recorded {row['identifier']} ({row['channel']})" if ok
          else f"[ledger] {row['identifier']} was already recorded — not sent twice")
    ledger.close()
    return 0 if ok else 1


def cmd_suppress(identifier: str, reason: str) -> int:
    ledger = Ledger()
    for channel in (config.CHANNEL_EMAIL, config.CHANNEL_INSTAGRAM):
        ledger.suppress(channel, identifier, reason)
    print(f"[ledger] {identifier} suppressed on all channels — never contacted again")
    ledger.close()
    return 0


def cmd_stats() -> int:
    ledger = Ledger()
    s = ledger.stats()
    print(f"sent: {s['sent']}  suppressed: {s['suppressed']}")
    for channel, count in sorted(s["by_channel"].items()):
        print(f"  {channel}: {count}")
    ledger.close()
    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Prepare cold outreach drafts. Never sends anything.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--prep", action="store_true",
                       help="build drafts.md + queue.xlsx from the scraper outputs")
    group.add_argument("--mark-sent", metavar="IDENT",
                       help="record an email/handle you just sent by hand")
    group.add_argument("--suppress", metavar="IDENT",
                       help="permanent opt-out across every channel")
    group.add_argument("--stats", action="store_true", help="show ledger totals")
    parser.add_argument("--limit", type=int, default=0,
                        help="cap how many drafts --prep produces (a day's batch)")
    parser.add_argument("--reason", default="replied stop",
                        help="reason recorded alongside --suppress")
    args = parser.parse_args()

    if args.prep:
        sys.exit(cmd_prep(args.limit))
    if args.mark_sent:
        sys.exit(cmd_mark_sent(args.mark_sent))
    if args.suppress:
        sys.exit(cmd_suppress(args.suppress, args.reason))
    sys.exit(cmd_stats())
