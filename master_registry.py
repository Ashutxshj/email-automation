"""The single source of truth: Projects/leads_master.xlsx.

One Excel file. Every scraper appends to it; /email-automation reads it, writes
the Message column, and flips Reached / Do Not Contact. Nothing else persists
lead state anywhere — no per-repo registry, no SQLite, no output/ folder.

This file is copied verbatim into /scraper, /scraper2, /scraper3 and
/email-automation. They are separate git repos with no shared package, so a copy
is the only way to share it. If you change one, change them all.

Dedup keys, strongest first: email, instagram handle, phone (last 10 digits).
The business name is used ONLY when a row has none of those — two branches of
"Looks Salon" share a name but not a phone, and collapsing them would lose a
real lead.
"""

import os
import re
from datetime import datetime

# Projects/leads_master.xlsx — one level up from whichever repo imports this.
_DEFAULT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "leads_master.xlsx")
MASTER_FILE = os.getenv("MASTER_FILE", _DEFAULT)

BULLETS_COLUMN = "Bullet points about this business which email automation can used"

COLUMNS = [
    "Business Name",
    "Category",
    "Lead Type",
    "First Reported At",
    "Has_Website",
    "Phone Number",
    "Email Address",
    "Instagram",
    "Rating",
    "Reviews",
    BULLETS_COLUMN,
    "Message",
    "Reached",
    "Do Not Contact",
]

# Written by a scraper, never by email-automation.
SOURCE_COLUMNS = COLUMNS[:11]
# Written by email-automation, never by a scraper.
OUTREACH_COLUMNS = ["Message", "Reached", "Do Not Contact"]

LEAD_TYPE_GOLDENROD = "Goldenrod"
LEAD_TYPE_NEW_BARK = "New Bark"
LEAD_TYPE_STALE_SITE = "Stale Website"


# --- identity ---------------------------------------------------------------

def norm_phone(phone) -> str:
    digits = re.sub(r"\D", "", str(phone or ""))
    return digits[-10:] if len(digits) >= 10 else digits


def norm_handle(handle) -> str:
    return str(handle or "").strip().lstrip("@").lower()


def norm_email(email) -> str:
    return str(email or "").strip().lower()


def identity_keys(row: dict) -> set[str]:
    """Strong keys if any exist; the name is a last resort, not an equal peer."""
    keys = set()
    email = norm_email(row.get("Email Address"))
    if email:
        keys.add(f"em:{email}")
    handle = norm_handle(row.get("Instagram"))
    if handle:
        keys.add(f"ig:{handle}")
    phone = norm_phone(row.get("Phone Number"))
    if phone:
        keys.add(f"ph:{phone}")
    if not keys:
        name = str(row.get("Business Name") or "").strip().lower()
        if name:
            keys.add(f"na:{name}")
    return keys


def matches(identifier: str, row: dict) -> bool:
    """Does this row correspond to the email address or @handle given on the CLI?"""
    ident = str(identifier or "").strip().lower()
    if not ident:
        return False
    if "@" in ident and not ident.startswith("@"):
        return norm_email(row.get("Email Address")) == ident
    return norm_handle(row.get("Instagram")) == norm_handle(ident)


# --- bullets ----------------------------------------------------------------

# Only objectively checkable observations. site_audit.py also emits opinions
# ("design/markup looks passable but long-neglected") and failure notes ("site
# could not be analyzed"); quoting one of those as fact is how you lose a lead,
# so they are simply never turned into a bullet.
_AUDIT_BULLETS = [
    ("no https", "Site loads over plain HTTP, so Chrome shows visitors a \"Not secure\" warning"),
    ("no mobile viewport", "No mobile viewport set — the site renders desktop-width on phones"),
    ("embeds flash", "Homepage still embeds Flash, which no browser has run since 2020"),
    ("obsolete 1990s-era html", "Homepage is still built with 1990s-era tags like <font> and <center>"),
    ("decade-old jquery", "Loads jQuery 1.x, roughly a decade out of date"),
    ("missing page title", "Homepage has no title tag — tabs and Google results show a bare URL"),
    ("missing meta description", "No meta description, so Google writes the search snippet itself"),
]
_COPYRIGHT_RE = re.compile(r"copyright stuck at (\d{4})", re.I)
_STALE_DAYS_RE = re.compile(r"\((\d+)\s+days ago", re.I)


def stale_days(last_updated: str) -> int | None:
    match = _STALE_DAYS_RE.search(str(last_updated or ""))
    return int(match.group(1)) if match else None


def website_bullets(issues: str, last_updated: str = "") -> str:
    """Why this business's existing site needs a redesign. Checkable facts only."""
    text = str(issues or "").lower()
    out = [clause for key, clause in _AUDIT_BULLETS if key in text]

    year = _COPYRIGHT_RE.search(str(issues or ""))
    if year:
        out.append(f"Footer copyright still reads {year.group(1)}")

    days = stale_days(last_updated)
    if days and days > 365:
        years = days // 365
        out.append(f"Nothing on the site has changed in about "
                   f"{years} year{'s' if years > 1 else ''} ({days} days)")
    return "\n".join(f"• {b}" for b in out)


def no_website_bullets(category: str, rating, reviews, instagram: str = "") -> str:
    """Why this business would benefit from having a website at all.

    The first bullet is the hook, and /email-automation drops it straight into a
    sentence ("I was looking up X and noticed <bullet>"). So it must read as a
    noun phrase, not a predicate: "4.9 stars from 133 reviews, but no website",
    never "Rated 4.9 stars..." — which would render as "noticed rated 4.9 stars".
    """
    out = []
    cat = (str(category or "business")).strip().lower() or "business"
    try:
        review_count = int(float(reviews or 0))
    except (TypeError, ValueError):
        review_count = 0

    if rating and review_count:
        out.append(f"{rating} stars from {review_count:,} Google reviews, "
                   "but no website to send that traffic to")
    elif review_count:
        out.append(f"{review_count:,} Google reviews, but no website")
    else:
        out.append("a Google Maps listing, but no website")

    out.append(f"People searching \"{cat} near me\" land on competitors who have a site")
    out.append("No way to show prices, timings or take bookings without one")
    if instagram:
        out.append("An Instagram page doesn't rank in Google search the way a site does")
    return "\n".join(f"• {b}" for b in out)


# --- workbook ---------------------------------------------------------------

def _open():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    if os.path.exists(MASTER_FILE):
        wb = load_workbook(MASTER_FILE)
        return wb, wb.active

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, col in enumerate(COLUMNS, 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 60 if col in (BULLETS_COLUMN, "Message") \
            else min(len(col) + 6, 30)
    ws.freeze_panes = "A2"
    ws["K2"].alignment = Alignment(wrap_text=True, vertical="top")
    return wb, ws


def _row_dict(header: list, values: tuple) -> dict:
    return {col: (values[i] if i < len(values) and values[i] is not None else "")
            for i, col in enumerate(header)}


def load_rows() -> list[dict]:
    """Every lead in the master file. [] if it doesn't exist yet."""
    if not os.path.exists(MASTER_FILE):
        return []
    from openpyxl import load_workbook
    wb = load_workbook(MASTER_FILE, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h) for h in (next(rows, None) or []) if h is not None]
    out = [_row_dict(header, values) for values in rows
           if any(v is not None for v in values)]
    wb.close()
    return out


def load_keys() -> set[str]:
    """Identity keys of every business already in the master file."""
    keys: set[str] = set()
    for row in load_rows():
        keys |= identity_keys(row)
    return keys


def upsert(rows: list[dict]) -> int:
    """Append rows that aren't already present. Returns how many were added.

    Never raises: a registry failure must not lose a completed scrape. The
    caller has already spent the API credits.
    """
    if not rows:
        return 0
    try:
        from openpyxl.styles import Alignment
        wb, ws = _open()
        header = [str(c.value) for c in ws[1]]
        existing: set[str] = set()
        for values in ws.iter_rows(min_row=2, values_only=True):
            existing |= identity_keys(_row_dict(header, values))

        stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        added = 0
        for row in rows:
            keys = identity_keys(row)
            if not keys or (keys & existing):
                continue          # already known, or nothing to identify it by
            existing |= keys
            record = dict(row)
            record.setdefault("First Reported At", stamp)
            record.setdefault("Message", "")
            record.setdefault("Reached", False)
            record.setdefault("Do Not Contact", False)
            ws.append([record.get(col, "") for col in header])
            for col in (BULLETS_COLUMN, "Message"):
                if col in header:
                    ws.cell(row=ws.max_row, column=header.index(col) + 1).alignment = \
                        Alignment(wrap_text=True, vertical="top")
            added += 1
        wb.save(MASTER_FILE)
        return added
    except Exception as exc:
        print(f"[master] WARN: could not update {MASTER_FILE}: {exc}")
        return 0


def set_fields(identifier: str, **updates) -> bool:
    """Update one lead's outreach columns in place. True if a row was found."""
    bad = set(updates) - set(OUTREACH_COLUMNS)
    if bad:
        raise ValueError(f"{sorted(bad)} are scraper-owned columns, not outreach ones")
    if not os.path.exists(MASTER_FILE):
        return False

    from openpyxl.styles import Alignment
    wb, ws = _open()
    header = [str(c.value) for c in ws[1]]
    for excel_row in range(2, ws.max_row + 1):
        values = tuple(ws.cell(row=excel_row, column=i + 1).value
                       for i in range(len(header)))
        if not matches(identifier, _row_dict(header, values)):
            continue
        for col, value in updates.items():
            cell = ws.cell(row=excel_row, column=header.index(col) + 1)
            cell.value = value
            if col == "Message":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        wb.save(MASTER_FILE)
        return True
    return False


def is_true(value) -> bool:
    """Excel round-trips booleans as bool, 'TRUE', 1, or ''. Normalise."""
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "1", "yes")
